"""
Stock sheet search + update logic.
Target: utils/sheets.py
"""

import datetime
import io

import openpyxl
import pytest

from utils import sheets as sheets_mod


# ---------------------------------------------------------------------------
# get_active_sheet_folder
# ---------------------------------------------------------------------------

def test_active_folder_found_by_substring(seeded_drive, drive_id):
    folder_id, folder_name = sheets_mod.get_active_sheet_folder(
        seeded_drive, drive_id
    )
    assert folder_id == seeded_drive.ids["active"]
    assert "Currently in use" in folder_name


def test_active_folder_raises_when_stock_sheets_missing(fake_drive, drive_id):
    with pytest.raises(FileNotFoundError, match="Stock Sheets"):
        sheets_mod.get_active_sheet_folder(fake_drive, drive_id)


def test_active_folder_raises_when_no_currently_in_use(fake_drive, drive_id):
    fake_drive.add_folder("Stock Sheets", drive_id)  # no active subfolder
    with pytest.raises(FileNotFoundError, match="Currently in use"):
        sheets_mod.get_active_sheet_folder(fake_drive, drive_id)


def test_active_folder_name_never_hardcoded(seeded_drive, drive_id):
    """
    Rename the active folder to something arbitrary but still containing
    'Currently in use' — search must still succeed.
    """
    svc = seeded_drive
    svc.records[svc.ids["active"]]["name"] = "ZZ_Q2_2026_Currently in use_legacy"
    fid, name = sheets_mod.get_active_sheet_folder(svc, drive_id)
    assert fid == svc.ids["active"]
    assert "Currently in use" in name


# ---------------------------------------------------------------------------
# list_serial_number_sheets
# ---------------------------------------------------------------------------

def test_list_serial_sheets_includes_only_serial_number_listing(
    seeded_drive, drive_id
):
    svc = seeded_drive
    active = svc.ids["active"]
    a = svc.add_file(
        "Serial Number Listing CPT.xlsx", active,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    b = svc.add_file(
        "Serial Number Listing FMAS Digital Trio.xlsx", active,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    # Inventory Levels file must NOT be returned — name filter excludes it.
    svc.add_file("Inventory Levels CPT.xlsx", active)

    result = sheets_mod.list_serial_number_sheets(svc, active, drive_id)
    ids = {r["id"] for r in result}
    assert ids == {a, b}


# ---------------------------------------------------------------------------
# find_serial_number — happy paths + aborts
# ---------------------------------------------------------------------------

def _seed_one_sheet(drive, sheet_bytes):
    file_id = drive.add_file(
        "Serial Number Listing CPT.xlsx",
        drive.ids["active"],
        content=sheet_bytes,
    )
    drive.content[file_id] = sheet_bytes
    return file_id


def _patch_download(monkeypatch, drive):
    """Redirect _download_xlsx to read from FakeDrive.content."""
    def fake_download(service, file_id):
        buf = io.BytesIO(drive.content[file_id])
        return openpyxl.load_workbook(buf)
    monkeypatch.setattr(sheets_mod, "_download_xlsx", fake_download)


def test_find_serial_returns_row_details(
    seeded_drive, drive_id, sample_sheet_bytes, monkeypatch
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)

    result = sheets_mod.find_serial_number(seeded_drive, drive_id, "SN-0001")
    assert result is not None
    assert result["file_name"] == "Serial Number Listing CPT.xlsx"
    assert result["sheet_name"] == "Stock"
    assert result["row_values"][0] == "SN-0001"
    assert "Serial Number" in [str(h) for h in result["headers"]]


def test_find_serial_case_insensitive(
    seeded_drive, sample_sheet_bytes, monkeypatch, drive_id
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    assert sheets_mod.find_serial_number(seeded_drive, drive_id, "sn-0001") is not None


def test_find_serial_matches_numeric_stored_value(
    seeded_drive, sample_sheet_bytes, monkeypatch, drive_id
):
    """Excel may store a numeric serial as int — search must still match."""
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    result = sheets_mod.find_serial_number(
        seeded_drive, drive_id, "200254233608"
    )
    assert result is not None
    assert result["row_values"][0] == 200254233608


def test_find_serial_searches_all_sheets_before_giving_up(
    seeded_drive, sample_sheet_bytes, other_sheet_bytes, monkeypatch, drive_id
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    seeded_drive.add_file(
        "Serial Number Listing FMAS.xlsx",
        seeded_drive.ids["active"],
        content=other_sheet_bytes,
    )
    _patch_download(monkeypatch, seeded_drive)

    # SN-9999 is only in the second sheet
    result = sheets_mod.find_serial_number(seeded_drive, drive_id, "SN-9999")
    assert result is not None
    assert result["file_name"] == "Serial Number Listing FMAS.xlsx"


def test_find_serial_returns_none_when_missing(
    seeded_drive, sample_sheet_bytes, monkeypatch, drive_id
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    assert sheets_mod.find_serial_number(seeded_drive, drive_id, "NOT-HERE") is None


def test_find_serial_raises_if_no_sheets_present(seeded_drive, drive_id):
    # Active folder exists but contains no Serial Number Listing files.
    with pytest.raises(FileNotFoundError, match="Serial Number Listing"):
        sheets_mod.find_serial_number(seeded_drive, drive_id, "SN-0001")


# ---------------------------------------------------------------------------
# update_stock_row — writes + highlight + abort-on-missing
# ---------------------------------------------------------------------------

def _capture_uploaded_workbook(monkeypatch):
    """Patch _upload_xlsx so we can inspect the written workbook."""
    captured = {}

    def fake_upload(service, file_id, workbook):
        buf = io.BytesIO()
        workbook.save(buf)
        buf.seek(0)
        captured["file_id"] = file_id
        captured["wb"] = openpyxl.load_workbook(buf)

    monkeypatch.setattr(sheets_mod, "_upload_xlsx", fake_upload)
    return captured


def test_update_stock_row_writes_account_date_and_red_fill(
    seeded_drive, drive_id, sample_sheet_bytes, monkeypatch
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    captured = _capture_uploaded_workbook(monkeypatch)

    sheets_mod.update_stock_row(
        seeded_drive, drive_id, "SN-0001", "Unit 42 Atlantic Beach"
    )

    ws = captured["wb"]["Stock"]
    # Row 3 in the fixture: first data row after title + headers
    assert ws.cell(row=3, column=1).value == "SN-0001"
    assert ws.cell(row=3, column=3).value == "Unit 42 Atlantic Beach"
    today_iso = datetime.date.today().strftime("%Y-%m-%d")
    assert ws.cell(row=3, column=4).value == today_iso

    # Entire row should be red.
    for col in range(1, ws.max_column + 1):
        fill = ws.cell(row=3, column=col).fill
        # openpyxl stores colour as ARGB; FF0000 is the pure-red we set.
        assert fill.fill_type == "solid"
        assert str(fill.start_color.rgb).upper().endswith("FF0000")


def test_update_stock_row_raises_without_writing_when_serial_missing(
    seeded_drive, drive_id, sample_sheet_bytes, monkeypatch
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    upload_called = {"n": 0}

    def fake_upload(*a, **kw):
        upload_called["n"] += 1

    monkeypatch.setattr(sheets_mod, "_upload_xlsx", fake_upload)

    with pytest.raises(ValueError, match="not found"):
        sheets_mod.update_stock_row(
            seeded_drive, drive_id, "DOES-NOT-EXIST", "Unit 1 Site"
        )
    assert upload_called["n"] == 0


def test_update_stock_row_does_not_touch_other_rows(
    seeded_drive, drive_id, sample_sheet_bytes, monkeypatch
):
    _seed_one_sheet(seeded_drive, sample_sheet_bytes)
    _patch_download(monkeypatch, seeded_drive)
    captured = _capture_uploaded_workbook(monkeypatch)

    sheets_mod.update_stock_row(
        seeded_drive, drive_id, "SN-0001", "Unit 42 Atlantic Beach"
    )

    ws = captured["wb"]["Stock"]
    # Untouched row 4 should retain original values and no red fill.
    assert ws.cell(row=4, column=1).value == "SN-0002"
    assert ws.cell(row=4, column=3).value == "Unit 5 Atlantic Beach"
    fill = ws.cell(row=4, column=1).fill
    rgb = str(fill.start_color.rgb).upper() if fill.start_color else ""
    assert not rgb.endswith("FF0000")
