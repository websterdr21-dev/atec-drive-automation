"""
Sheet operations for Serial Number Listing .xlsx files stored in Google Drive.

Flow:
  1. Find "Stock Sheets" folder in Shared Drive root
  2. Find the single subfolder whose name contains "Currently in use"
  3. List all files whose name starts with "Serial Number Listing"
  4. Download each .xlsx, search with openpyxl
"""

import io
import os
import datetime
import pathlib
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl
from openpyxl.styles import PatternFill
from utils.auth import get_drive_service


RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

_CACHE_DIR = pathlib.Path(__file__).parent.parent / "data" / "sheet_cache"


# ---------------------------------------------------------------------------
# Folder helpers
# ---------------------------------------------------------------------------

def _find_folder(service, name_exact=None, name_contains=None, parent_id=None, drive_id=None):
    """Return the first folder matching name criteria under parent_id."""
    parts = ["mimeType='application/vnd.google-apps.folder'", "trashed=false"]
    if name_exact:
        parts.append(f"name='{name_exact}'")
    if name_contains:
        parts.append(f"name contains '{name_contains}'")
    if parent_id:
        parts.append(f"'{parent_id}' in parents")

    results = service.files().list(
        q=" and ".join(parts),
        corpora="drive",
        driveId=drive_id,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name)",
    ).execute()

    files = results.get("files", [])
    return files[0] if files else None


def get_active_sheet_folder(service, drive_id):
    """
    Returns (folder_id, folder_name) of the active 'Currently in use' subfolder
    inside 'Stock Sheets'.
    Raises if either folder is not found.
    """
    stock_sheets = _find_folder(service, name_exact="Stock Sheets", drive_id=drive_id)
    if not stock_sheets:
        raise FileNotFoundError("'Stock Sheets' folder not found in Shared Drive root.")

    active = _find_folder(
        service,
        name_contains="Currently in use",
        parent_id=stock_sheets["id"],
        drive_id=drive_id,
    )
    if not active:
        raise FileNotFoundError(
            "No subfolder containing 'Currently in use' found inside 'Stock Sheets'."
        )

    return active["id"], active["name"]


def list_serial_number_sheets(service, folder_id, drive_id):
    """
    Returns list of {id, name} dicts for all 'Serial Number Listing' .xlsx files
    in the given folder. Ignores 'Inventory Levels' files.
    """
    results = service.files().list(
        q=f"'{folder_id}' in parents and name contains 'Serial Number Listing' and trashed=false",
        corpora="drive",
        driveId=drive_id,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        fields="files(id, name, modifiedTime)",
    ).execute()
    return results.get("files", [])


# ---------------------------------------------------------------------------
# Download / upload helpers
# ---------------------------------------------------------------------------

def _download_xlsx(service, file_id):
    """Download a Drive file and return an openpyxl Workbook."""
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return openpyxl.load_workbook(buf)


def _download_xlsx_cached(service, file_id, modified_time):
    """Return cached workbook if Drive modifiedTime matches; download and cache otherwise.

    Delegates to _download_xlsx on cache miss so tests can patch _download_xlsx normally.
    """
    if modified_time:
        meta = _CACHE_DIR / f"{file_id}.meta"
        cached = _CACHE_DIR / f"{file_id}.xlsx"
        if meta.exists() and cached.exists() and meta.read_text().strip() == modified_time:
            return openpyxl.load_workbook(cached)

    wb = _download_xlsx(service, file_id)

    if modified_time:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        buf = io.BytesIO()
        wb.save(buf)
        (_CACHE_DIR / f"{file_id}.xlsx").write_bytes(buf.getvalue())
        (_CACHE_DIR / f"{file_id}.meta").write_text(modified_time)

    return wb


def _upload_xlsx(service, file_id, workbook):
    """Upload a modified openpyxl Workbook back to the same Drive file."""
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    service.files().update(
        fileId=file_id,
        media_body=media,
        supportsAllDrives=True,
    ).execute()


# ---------------------------------------------------------------------------
# Core search
# ---------------------------------------------------------------------------

def find_serial_number(service, drive_id, serial_number):
    """
    Search all Serial Number Listing sheets for serial_number.

    Returns:
        {
            "file_id": str,
            "file_name": str,
            "sheet_name": str,
            "row_index": int,          # 1-based
            "row_values": list,
            "headers": list,
        }
    or None if not found.
    """
    folder_id, folder_name = get_active_sheet_folder(service, drive_id)
    sheets = list_serial_number_sheets(service, folder_id, drive_id)

    if not sheets:
        raise FileNotFoundError("No 'Serial Number Listing' files found in the active folder.")

    serial_str = serial_number.strip()
    serial_lower = serial_str.lower()
    # Also prepare a numeric form for serials stored as int/float in Excel
    # e.g. user enters "0200254233608" → stored as 200254233608 or 200254233608.0
    try:
        serial_int = int(serial_str)
    except ValueError:
        serial_int = None

    def _matches(cell_value):
        if cell_value is None:
            return False
        # String comparison (case-insensitive)
        if str(cell_value).strip().lower() == serial_lower:
            return True
        # Numeric comparison: handle int/float stored values
        if serial_int is not None:
            if isinstance(cell_value, (int, float)) and int(cell_value) == serial_int:
                return True
        return False

    for file_info in sheets:
        wb = _download_xlsx_cached(service, file_info["id"], file_info.get("modifiedTime"))
        for ws in wb.worksheets:
            # Find the real header row (first row where col A is "Serial Number")
            header_row_idx = None
            headers = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row and str(row[0]).strip().lower() == "serial number":
                    header_row_idx = r_idx
                    headers = list(row)
                    break

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if header_row_idx and row_idx <= header_row_idx:
                    continue  # skip title/header rows
                for cell_value in row:
                    if _matches(cell_value):
                        return {
                            "file_id": file_info["id"],
                            "file_name": file_info["name"],
                            "sheet_name": ws.title,
                            "row_index": row_idx,
                            "row_values": list(row),
                            "headers": headers,
                        }
    return None


# ---------------------------------------------------------------------------
# Sheet update
# ---------------------------------------------------------------------------

def update_stock_row(service, drive_id, serial_number, full_address):
    """
    Find serial_number, update Current Account + Date Last Move, highlight row RED.
    Returns the result dict from find_serial_number on success.
    Raises ValueError if serial number not found.
    """
    result = find_serial_number(service, drive_id, serial_number)
    if result is None:
        raise ValueError(f"Serial number '{serial_number}' not found in any Serial Number Listing sheet.")

    folder_id, _ = get_active_sheet_folder(service, drive_id)
    wb = _download_xlsx(service, result["file_id"])
    ws = wb[result["sheet_name"]]

    headers = result["headers"]
    row_idx = result["row_index"]

    # Find column indices (case-insensitive)
    headers_lower = [str(h).strip().lower() if h else "" for h in headers]
    account_col = next((i + 1 for i, h in enumerate(headers_lower) if "current account" in h), None)
    date_col = next((i + 1 for i, h in enumerate(headers_lower) if "date last move" in h), None)

    if account_col:
        ws.cell(row=row_idx, column=account_col).value = full_address
    if date_col:
        ws.cell(row=row_idx, column=date_col).value = datetime.date.today().strftime("%Y-%m-%d")

    # Highlight entire row red
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col).fill = RED_FILL

    _upload_xlsx(service, result["file_id"], wb)
    return result
